# -*- coding: utf-8 -*-
import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

r115 = json.load(open('r115q.json', encoding='utf-8'))
FIVE = {'國':{'頂':13,'前':12,'均':10,'後':9,'底':7},'英':{'頂':13,'前':11,'均':8,'後':5,'底':3},'數A':{'頂':12,'前':10,'均':8,'後':5,'底':4},'數B':{'頂':11,'前':9,'均':5,'後':3,'底':2},'社':{'頂':13,'前':12,'均':10,'後':8,'底':7},'自':{'頂':13,'前':12,'均':9,'後':7,'底':5}}
G15CELL = {'國':'$D$2','英':'$D$3','數A':'$D$4','數B':'$D$5','自':'$D$6','社':'$D$7'}
hist = json.load(open('hist.json', encoding='utf-8'))

def nk(school, dept):
    return school + '|' + re.sub(r'[()（）\s]', '', dept)

hist2 = {y: {nk(v['school'], v['dept']): v for v in hist[y].values()} for y in hist}

SUBJECTS = ['國','英','數A','數B','自','社','數甲','數乙','物','化','生','歷','地','公','術']
SUBJ_LABEL = {'國':'國文(學測)','英':'英文(學測)','數A':'數學A(學測)','數B':'數學B(學測)',
              '自':'自然(學測)','社':'社會(學測)','數甲':'數學甲(分科)','數乙':'數學乙(分科)',
              '物':'物理(分科)','化':'化學(分科)','生':'生物(分科)','歷':'歷史(分科)',
              '地':'地理(分科)','公':'公民與社會(分科)','術':'術科(百分制)'}

def subj_text(d):
    return ' '.join(f'{s}x{d[s]:.2f}' for s in SUBJECTS if s in d)

ZH = 'Microsoft JhengHei'
f_hdr = Font(name=ZH, size=10, bold=True, color='FFFFFF')
f_body = Font(name=ZH, size=10)
f_note = Font(name=ZH, size=10, color='C00000')
f_green = Font(name=ZH, size=10, color='008000')
f_blue = Font(name=ZH, size=11, color='0000FF')
f_bold = Font(name=ZH, size=10, bold=True)
fill_hdr = PatternFill('solid', fgColor='1F4E79')
fill_hdr2 = PatternFill('solid', fgColor='548235')
fill_in = PatternFill('solid', fgColor='FFFF00')
fill_alt = PatternFill('solid', fgColor='DDEBF7')
thin = Border(*[Side(style='thin', color='BFBFBF')]*4)
ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
lft = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = Workbook()

# ---------- 說明 ----------
ws0 = wb.active; ws0.title = '說明'
ws0.column_dimensions['A'].width = 4
ws0.column_dimensions['B'].width = 112
rows = [
 ('115學年度大學分發入學 採計總分試算、114差異分析與錄取機會評估（v5通用版（預設空白）學測級分自動換算＋五標/英聽檢定自動判定）', True),
 ('', False),
 ('【使用方式】', True),
 ('1. 於「成績輸入」工作表黃色儲存格填入各科成績：學測與分科測驗均填入大考中心公告之「60級分制」成績（0～60），術科填百分制成績（0～100）。', False),
 ('2. 「校系總分計算」工作表自動計算：採計總分（SUMPRODUCT）、與114最低錄取分之差異、錄取機會評估。', False),
 ('', False),
 ('【錄取機會評估分級】（差異比率＝(採計總分－114最低錄取分)÷114最低錄取分）', True),
 ('・高：+10% 以上　・中高：+3%～+10%　・邊緣(五五波)：-3%～+3%　・偏低：-10%～-3%　・低：低於 -10%', False),
 ('・「114採計一致」欄為「否」者，評估結果附註「僅供趨勢參考」；為「無」者代表114無同名系組，顯示「無114分數可比」。', False),
 ('', False),
 ('【重要注意事項】', True),
 ('1. 錄取機會評估為「與114年靜態比較」，未反映115年考題難易、級分分布、名額增減與登記人數變化，非真實錄取機率。', False),
 ('2. 依115簡章：學測及分科測驗採計時成績均以60級分制計算，術科以百分制計算；60級分制成績查詢方式詳見大考中心115學年度考試簡章。', False),
 ('3. 114學年度為分科測驗首次加考數學乙；112/113年多數商管社科系組採「學測數B或數A」，114年起多改採「分科數乙」，備註欄逐系組標示。', False),
 ('4. 「檢定標準」為學測及英聽門檻，未達檢定者不予分發。採計科目未報考者不予分發；報考但缺考該科以0級分計。', False),
 ('5. 歷年最低錄取分數為該年度採計方式下之加權總分；採計不同年度之分數不可與本表試算總分直接比較。', False),
 ('', False),
 ('【資料來源與核實紀錄】', True),
 ('・115採計：115學年度大學分發入學招生簡章校系分則（PDF第24～266頁），共 %d 系組；以「字詞座標」與「字元座標」雙引擎независ解析交叉比對。' % len(r115), False),
 ('・臺大/政大/清華/陽明交通/成大五校共255系組：雙引擎解析結果100%%一致，並抽樣頁面以影像目視逐列核對（台大p24、政大p55）。', False),
 ('・112/113/114最低錄取分數：考分會各年度「各系組最低錄取標準及錄取人數一覽表」；五校已以原始榜單影像逐列目視核對（114：台大2頁＋政大＋清大＋陽明交大＋成大各1頁；113/112：台大第1頁），另全量執行異常掃描（有錄取無分數、無錄取有分數、分數範圍外均為0筆）。', False),
 ('・歷年與115系組以「校名＋系組名」（忽略括號）比對；新設/更名/分組系組不強制對應。', False),
]
for i, (txt, bold) in enumerate(rows, start=1):
    c = ws0.cell(row=i, column=2, value=txt.replace('независ',''))
    c.font = Font(name=ZH, size=12 if i == 1 else 10, bold=bold)
    c.alignment = lft

# ---------- 成績輸入 ----------
ws1 = wb.create_sheet('成績輸入')
for col, w in zip('ABCDE', (12, 20, 16, 16, 58)): ws1.column_dimensions[col].width = w
for j, h in enumerate(['類別','科目','成績（60級分制）','學測級分(自動換算)','說明'], 1):
    c = ws1.cell(row=1, column=j, value=h)
    c.font = f_hdr; c.fill = fill_hdr; c.alignment = ctr; c.border = thin
cat = {'國':'學測','英':'學測','數A':'學測','數B':'學測','自':'學測','社':'學測',
       '數甲':'分科','數乙':'分科','物':'分科','化':'分科','生':'分科','歷':'分科','地':'分科','公':'分科','術':'術科'}
example = {s:0 for s in ['國','英','數A','數B','自','社','數甲','數乙','物','化','生','歷','地','公','術']}
g15def  = {'國':14,'英':10,'數A':12,'數B':12,'自':0,'社':14}
for i, s in enumerate(SUBJECTS, start=2):
    ws1.cell(row=i, column=1, value=cat[s]).font = f_body
    ws1.cell(row=i, column=2, value=SUBJ_LABEL[s]).font = f_body
    c = ws1.cell(row=i, column=3, value=example[s])
    c.font = f_blue; c.fill = fill_in; c.border = thin; c.alignment = ctr
    if s in g15def:
        cg = ws1.cell(row=i, column=4, value=f'=IF(C{i}=0,0,CEILING(C{i}/4,1))')
        cg.font = f_green; cg.border = thin; cg.alignment = ctr
    for j in (1, 2): ws1.cell(row=i, column=j).border = thin
# 英聽級數輸入（第17列）
ws1.cell(row=17, column=1, value='英聽').font = f_body
ws1.cell(row=17, column=2, value='高中英語聽力測驗').font = f_body
ce = ws1.cell(row=17, column=3, value='未報考')
ce.font = f_blue; ce.fill = fill_in; ce.border = thin; ce.alignment = ctr
for j in (1,2): ws1.cell(row=17, column=j).border = thin
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type='list', formula1='"A級,B級,C級,F級,未報考"', allow_blank=False)
dv.error = '請由下拉選單選擇：A級、B級、C級、F級、未報考'
ws1.add_data_validation(dv); dv.add(ce)
note = ws1.cell(row=2, column=5, value='黃色儲存格為輸入區，目前為「範例成績」，請全部替換為實際成績。'
    '學測與分科均填大考中心公告之60級分制成績(0~60)；術科填百分制(0~100)；未報考之科目請填0。'
    '【請先輸入成績】所有黃色儲存格目前為 0／未報考，請填入您的成績後本表才會產生有效結果：學測與分科各科填大考中心公告之60級分制成績（0~60，術科填百分制0~100），未報考科目保留0；英聽以下拉選單選擇級數。「學測級分」欄由60級分自動換算（15級分=無條件進位(60級分÷4)，綠字勿手動修改），供210個系組之學測檢定自動判定（依115學測五標）；英聽檢定20系組亦自動判定；檢定未通過者不予分發。')
note.font = f_note; note.alignment = lft
ws1.merge_cells(start_row=2, start_column=5, end_row=17, end_column=5)

# ---------- 校系總分計算 ----------
ws = wb.create_sheet('校系總分計算')
NW = len(SUBJECTS); c0 = 9
cw_end = c0 + NW - 1              # T
COLS = (['115系組代碼','校名','系組名','學測及英聽檢定標準','英聽要求','英聽檢定','學測檢定','同分參酌順序'] + SUBJECTS +
        ['採計科目及加權(115)','採計總分(試算)','114最低錄取分','與114差異','114採計一致','錄取機會評估(相對114)',
         '113最低錄取分','112最低錄取分','114採計','113採計','112採計','備註(歷年採計差異)'])
col_txt   = cw_end + 1   # U
col_tot   = cw_end + 2   # V
col_114   = cw_end + 3   # W
col_diff  = cw_end + 4   # X
col_same  = cw_end + 5   # Y
col_eval  = cw_end + 6   # Z
col_113   = cw_end + 7   # AA
col_112   = cw_end + 8   # AB
col_a114  = cw_end + 9   # AC
col_a113  = cw_end + 10  # AD
col_a112  = cw_end + 11  # AE
col_note  = cw_end + 12  # AF

lab = ws.cell(row=1, column=c0-1, value='目前輸入成績→')
lab.font = Font(name=ZH, size=9, bold=True, color='008000'); lab.alignment = Alignment(horizontal='right')
for i in range(NW):
    c = ws.cell(row=1, column=c0+i, value=f'=成績輸入!$C${2+i}')
    c.font = f_green; c.alignment = ctr; c.number_format = '0'
for j, h in enumerate(COLS, 1):
    c = ws.cell(row=2, column=j, value=h)
    c.font = f_hdr
    c.fill = fill_hdr2 if j in (col_114, col_diff, col_same, col_eval) else fill_hdr
    c.alignment = ctr; c.border = thin
ws.row_dimensions[2].height = 32

cL0, cL1 = get_column_letter(c0), get_column_letter(cw_end)
W_, V_, Y_ = get_column_letter(col_114), get_column_letter(col_tot), get_column_letter(col_same)

r = 3
for rec in r115:
    d115 = {s: w for s, w, o in rec['subjects']}
    k = nk(rec['school'], rec['dept'])
    ws.cell(row=r, column=1, value=rec['code'])
    ws.cell(row=r, column=2, value=rec['school'])
    ws.cell(row=r, column=3, value=rec['dept'])
    ws.cell(row=r, column=4, value=rec['check'])
    m_e = re.search(r'英\s*聽\s*[（(]?\s*([ABＡＢ])\s*級', rec['check'])
    e_req = ({'Ａ':'A','Ｂ':'B'}.get(m_e.group(1), m_e.group(1)) + '級') if m_e else ''
    if e_req:
        ec = ws.cell(row=r, column=5, value=e_req); ec.font = f_bold; ec.alignment = ctr
        ej = ws.cell(row=r, column=6, value=(
            f'=IF(IF(成績輸入!$C$17="A級",4,IF(成績輸入!$C$17="B級",3,IF(成績輸入!$C$17="C級",2,IF(成績輸入!$C$17="F級",1,0))))>='
            f'IF(E{r}="A級",4,3),"通過","未通過")'))
        ej.alignment = ctr
    else:
        ws.cell(row=r, column=6, value='－').alignment = ctr
    if rec.get('q'):
        conds = []
        for cl in rec['q']:
            alts = ['成績輸入!' + G15CELL[s] + '>=' + str(FIVE[s][d]) for s, d in cl]
            conds.append('OR(' + ','.join(alts) + ')' if len(alts) > 1 else alts[0])
        expr = 'AND(' + ','.join(conds) + ')' if len(conds) > 1 else conds[0]
        qc = ws.cell(row=r, column=7, value=f'=IF({expr},"通過","未通過")')
        qc.alignment = ctr
    else:
        ws.cell(row=r, column=7, value='－').alignment = ctr
    order = sorted(rec['subjects'], key=lambda t: (t[2] if t[2] else 9))
    ws.cell(row=r, column=8, value='→'.join(s for s, w, o in order))
    for i, s in enumerate(SUBJECTS):
        if s in d115:
            c = ws.cell(row=r, column=c0+i, value=d115[s]); c.number_format = '0.00'
    ws.cell(row=r, column=col_txt, value=subj_text(d115))
    tot = ws.cell(row=r, column=col_tot, value=f'=SUMPRODUCT({cL0}{r}:{cL1}{r},{cL0}$1:{cL1}$1)')
    tot.number_format = '0.00'; tot.font = f_bold
    notes = []
    yr_map = {'114': (col_114, col_a114), '113': (col_113, col_a113), '112': (col_112, col_a112)}
    same114 = '無'
    for y in ('114','113','112'):
        cs, ca = yr_map[y]
        h = hist2[y].get(k)
        if h is None:
            notes.append(f'{y}年無同名系組')
            if y == '114': same114 = '無'
            continue
        if h['score']:
            sc = ws.cell(row=r, column=cs, value=float(h['score'])); sc.number_format = '0.00'
        else:
            ws.cell(row=r, column=cs, value='當年無錄取')
        ws.cell(row=r, column=ca, value=subj_text(h['subjects']))
        if h['subjects'] != d115:
            notes.append(f"{y}年採計不同({subj_text(h['subjects'])})")
            if y == '114': same114 = '否'
        elif y == '114':
            same114 = '是'
    sm = ws.cell(row=r, column=col_same, value=same114)
    sm.alignment = ctr
    sm.font = f_note if same114 != '是' else f_body
    df = ws.cell(row=r, column=col_diff,
                 value=f'=IF(ISNUMBER({W_}{r}),{V_}{r}-{W_}{r},"")')
    df.number_format = '+0.00;-0.00;0.00'; df.font = f_bold
    ev = ws.cell(row=r, column=col_eval, value=(
        f'=IF(NOT(ISNUMBER({W_}{r})),"無114分數可比",'
        f'IF(({V_}{r}-{W_}{r})/{W_}{r}>=0.1,"高",'
        f'IF(({V_}{r}-{W_}{r})/{W_}{r}>=0.03,"中高",'
        f'IF(({V_}{r}-{W_}{r})/{W_}{r}>=-0.03,"邊緣(五五波)",'
        f'IF(({V_}{r}-{W_}{r})/{W_}{r}>=-0.1,"偏低","低"))))'
        f'&IF({Y_}{r}="否","；114採計不同,僅供趨勢參考",""))'))
    ev.alignment = ctr
    if notes:
        nc = ws.cell(row=r, column=col_note, value='；'.join(notes)); nc.font = f_note
    for j in range(1, col_note+1):
        cell = ws.cell(row=r, column=j)
        cell.border = thin
        if cell.font is None or cell.font.name != ZH:
            cell.font = f_body
        wrap = j in (3, 4, col_txt, col_a114, col_a113, col_a112, col_note)
        cell.alignment = Alignment(vertical='center', wrap_text=wrap,
                                   horizontal='center' if j in (col_same, col_eval) else None)
        if r % 2 == 1:
            cell.fill = fill_alt
    r += 1

last = r - 1
ws.auto_filter.ref = f'A2:{get_column_letter(col_note)}{last}'
ws.freeze_panes = 'D3'
dcol = get_column_letter(col_diff)
ws.conditional_formatting.add(f'{dcol}3:{dcol}{last}',
    ColorScaleRule(start_type='num', start_value=-60, start_color='F8696B',
                   mid_type='num', mid_value=0, mid_color='FFEB84',
                   end_type='num', end_value=60, end_color='63BE7B'))
widths = {1:11, 2:16, 3:30, 4:22, 5:9, 6:9, 7:9, 8:15, col_txt:30, col_tot:11, col_114:11, col_diff:10,
          col_same:9, col_eval:22, col_113:11, col_112:11,
          col_a114:27, col_a113:27, col_a112:27, col_note:42}
for i in range(NW): widths[c0+i] = 6.5
for j, w in widths.items(): ws.column_dimensions[get_column_letter(j)].width = w

wb.save('115分科分發_採計總分試算與歷年比對_v5_通用版.xlsx')
print('rows:', last-2)
