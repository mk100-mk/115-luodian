# -*- coding: utf-8 -*-
"""
tools/parse_quota_reflow.py   （新增於 116學年度升級，2026-08-17）

計算各系組的「回流名額」：簡章核定分發名額 vs 回流後實際分發名額。

────────────────────────────────────────────────────────────
為什麼要做這件事
────────────────────────────────────────────────────────────
CLAUDE.md 鐵則：**分發名額 ≠ 申請名額，兩者性質不同。**

    分發名額（實際） = 教育部核定之登記分發名額 ＋ 申請入學未招滿之缺額回流

回流比例因校因系差異極大。只看「今年名額比去年多／少」就下判斷，
會把「回流造成的波動」誤讀為「系所擴招／縮編」，是名額趨勢分析最常見的錯誤。

────────────────────────────────────────────────────────────
資料來源
────────────────────────────────────────────────────────────
A. 核定名額：教育部【表7-2】115學年度日間學士班各院、系(組)、學位學程
   新生招生名額分配表（公告版 2026/05/06）
   → tools/115/quota_alloc_115.txt（TSV，欄位見檔頭）
   本腳本亦支援直接讀取原始 .pdf / .xlsx（若日後放入 tools/115/）。

B. 回流後名額：大學考試入學分發委員會「回流後分發入學總名額」
   → tools/115/count-115.xlsx（＝坊間流通之「招生名額115.xlsx」，MD5 相同）

────────────────────────────────────────────────────────────
兩份資料的層級不同（本腳本最關鍵的處理）
────────────────────────────────────────────────────────────
教育部核定表是「系」層級，分發榜單是「分組」層級。例如：

    核定表：國立臺灣大學 戲劇學系              33名（分發18）
    榜單  ：國立臺灣大學 戲劇學系(男) / 戲劇學系(女)   各7名

    核定表：銘傳大學(臺北) 企業管理學系
    榜單  ：銘傳大學 企業管理學系品牌行銷組(台北校區) 等 4 個分組

因此不能一對一比對。本腳本以「最長前綴比對」把榜單分組**向上彙總**到核定表的
系層級後才相減，並在輸出中保留 children 清單供追溯。
校區資訊（臺北／桃園／高雄）由榜單系組名的括號後綴取出，與核定表校名的
括號後綴對應。

────────────────────────────────────────────────────────────
轉錄完整性校驗（重要）
────────────────────────────────────────────────────────────
quota_alloc_115.txt 每校保留官方「總計」列。本腳本啟動時先逐校驗證
「各系組加總 == 官方總計」（5 個數值欄全部比對）。任何一欄對不上即中止執行，
不會帶著壞資料往下跑。這是為了讓轉錄錯誤變成**大聲的失敗**而非靜默的錯數字。

────────────────────────────────────────────────────────────
輸出 reflow.json
────────────────────────────────────────────────────────────
{
  "115": {
    "{核定表校名}|{核定表系組名}": {
       'school', 'dept', 'campus',
       'y114_approved':  114學年度核定招生名額,
       'y115_total':     115學年度分配名額小計(含擴充名額),
       'star':           繁星推薦名額,
       'apply':          申請入學名額,
       'dist_approved':  登記分發入學核定名額,
       'dist_actual':    回流後分發入學總名額（榜單分組加總）,
       'reflow':         回流量 = dist_actual - dist_approved,
       'reflow_pct':     回流幅度% = reflow / dist_approved * 100,
       'apply_share':    申請入學占小計比例%,
       'dist_share':     分發入學占小計比例%（核定口徑）,
       'children':       [對應之榜單分組名稱, ...]
    }, ...
  },
  "child_index": { "{榜單校名}|{榜單系組名}": "{核定表校名}|{核定表系組名}" },
  "school_summary": { "{核定表校名}": {...同上彙總...} },
  "_meta": {...}
}

用法：於 tools/ 目錄下執行   python parse_quota_reflow.py
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
YEAR = '115'

ALLOC_TSV = BASE / YEAR / f'quota_alloc_{YEAR}.txt'
COUNT_XLSX = BASE / YEAR / f'count-{YEAR}.xlsx'

# 欄序經實證確認（詳見 quota_alloc_115.txt 檔頭）：第3欄為「登記分發入學」而非繁星。
# 官方PDF是堆疊式表頭，文字抽取順序不可靠，故以資料本身反推：
#   (1) 台大61系組回流後名額全部 >= 第3欄（平均+4.6，無一為負）→ 符合「缺額只回流進、不流出」
#   (2) 本表68校加總 分發20,380 / 繁星15,539 / 申請50,075，與全國各管道規模相符
#   (3) 北藝大第3欄全為0，該校確實不參加考試分發入學
NUM_COLS = ['y114_approved', 'y115_total', 'dist_approved', 'star', 'apply']

CAMPUS_NORM = {'台北': '臺北', '臺北': '臺北', '桃園': '桃園',
               '高雄': '高雄', '臺南': '臺南', '台南': '臺南'}


def norm_name(s):
    """系組名正規化：移除括號（含內容以外的符號）與空白，全形統一。"""
    s = str(s).strip()
    s = s.replace('（', '(').replace('）', ')')
    s = re.sub(r'\s+', '', s)
    return s


def strip_parens(s):
    """移除所有括號符號但**保留內容**：'化學系(化學組)' → '化學系化學組'。

    保留內容而非刪除，是因為官方兩份資料對同一分組一邊寫括號、一邊不寫，
    刪掉內容會讓不同分組被誤併為同一系組。
    """
    return re.sub(r'[()]', '', s)


def split_campus(dept):
    """從榜單系組名尾端取出校區：'企業管理學系品牌行銷組(台北校區)' → (核心名, '臺北')"""
    m = re.search(r'\(([^()]*?)校區\)\s*$', dept)
    if not m:
        return dept, None
    core = dept[:m.start()]
    return core, CAMPUS_NORM.get(m.group(1), m.group(1))


def split_school_campus(school):
    """從核定表校名取出校區：'銘傳大學(臺北)' → ('銘傳大學', '臺北')"""
    m = re.search(r'\(([^()]+)\)\s*$', school)
    if not m:
        return school, None
    return school[:m.start()], CAMPUS_NORM.get(m.group(1), m.group(1))


# ────────────────────────── 讀取核定名額 ──────────────────────────
def load_alloc():
    if not ALLOC_TSV.exists():
        raise FileNotFoundError(
            f'找不到 {ALLOC_TSV}\n'
            '請將教育部【表7-2】115學年度招生名額分配表轉為 TSV 放入該路徑，'
            '欄位：校名/系組名/114核定/115小計/繁星/申請/分發，每校保留「總計」列。')

    depts, totals = [], {}
    for ln in ALLOC_TSV.read_text(encoding='utf-8').splitlines():
        if not ln.strip() or ln.lstrip().startswith('#'):
            continue
        p = ln.split('\t')
        if len(p) != 7:
            raise ValueError(f'欄位數不符（應為7）：{ln!r}')
        school, dept = p[0].strip(), p[1].strip()
        vals = [0 if v.strip() in ('--', '') else int(v) for v in p[2:]]
        if dept == '總計':
            totals[school] = vals
        else:
            depts.append((school, dept, vals))

    # ── 逐校校驗：各系組加總 == 官方總計 ──
    agg = {}
    for school, _, vals in depts:
        cur = agg.setdefault(school, [0] * 5)
        for i, v in enumerate(vals):
            cur[i] += v
    errs = []
    for school, exp in totals.items():
        got = agg.get(school, [0] * 5)
        if got != exp:
            errs.append(f'  {school}: 加總 {got} ≠ 官方總計 {exp} '
                        f'（差 {[g - e for g, e in zip(got, exp)]}）')
    missing = set(agg) - set(totals)
    for school in sorted(missing):
        errs.append(f'  {school}: 缺少官方「總計」列，無法校驗')
    if errs:
        raise SystemExit('【轉錄校驗失敗】以下學校資料與官方總計不符，已中止：\n'
                         + '\n'.join(errs))

    print(f'✔ 轉錄校驗通過：{len(totals)} 校 / {len(depts)} 系組，'
          f'5 個數值欄逐校加總全部吻合官方總計')
    return depts, totals


# ────────────────────────── 讀取回流後名額 ──────────────────────────
def load_count():
    wb = openpyxl.load_workbook(COUNT_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        school, dept, code, seat = r[1], r[2], r[3], r[4]
        if not school or not dept or seat is None:
            continue
        school = str(school).strip()
        if school.endswith('合計'):
            continue
        rows.append((school, str(dept).strip(), str(code).strip(), int(seat)))
    return rows


# ────────────────────────── 主流程 ──────────────────────────
def main():
    depts, _ = load_alloc()
    count_rows = load_count()

    # 建立 (校名base, 校區) → [(正規化系組名, 原始鍵)] 索引，長名在前供最長前綴比對
    index = {}
    recs = {}
    for school, dept, vals in depts:
        base, campus = split_school_campus(school)
        key = f'{school}|{dept}'
        norm = strip_parens(norm_name(dept))
        index.setdefault((base, campus), []).append((norm, key))
        recs[key] = dict(zip(NUM_COLS, vals),
                         school=school, dept=dept, campus=campus,
                         dist_actual=0, children=[])
    for k in index:
        index[k].sort(key=lambda t: -len(t[0]))   # 最長前綴優先

    child_index = {}
    unmatched = []
    for school, dept, code, seat in count_rows:
        core, campus = split_campus(dept)
        norm = strip_parens(norm_name(core))
        cands = index.get((school, campus)) or index.get((school, None)) or []
        hit = next((key for nm, key in cands if norm.startswith(nm)), None)
        if hit is None and campus is not None:
            cands = index.get((school, None)) or []
            hit = next((key for nm, key in cands if norm.startswith(nm)), None)
        if hit is None:
            unmatched.append((school, dept, code, seat))
            continue
        recs[hit]['dist_actual'] += seat
        recs[hit]['children'].append(dept)
        child_index[f'{school}|{dept}'] = hit

    # ── 計算回流量 ──
    matched = 0
    for key, r in recs.items():
        if not r['children']:
            r['reflow'] = None
            r['reflow_pct'] = None
        else:
            matched += 1
            r['reflow'] = r['dist_actual'] - r['dist_approved']
            r['reflow_pct'] = (round(r['reflow'] / r['dist_approved'] * 100, 1)
                               if r['dist_approved'] else None)
        tot = r['y115_total']
        r['apply_share'] = round(r['apply'] / tot * 100, 1) if tot else None
        r['dist_share'] = round(r['dist_approved'] / tot * 100, 1) if tot else None

    # ── 逐校彙總 ──
    school_summary = {}
    for r in recs.values():
        if not r['children']:
            continue
        s = school_summary.setdefault(r['school'], {
            'dist_approved': 0, 'dist_actual': 0, 'y115_total': 0,
            'star': 0, 'apply': 0, 'depts': 0})
        s['dist_approved'] += r['dist_approved']
        s['dist_actual'] += r['dist_actual']
        s['y115_total'] += r['y115_total']
        s['star'] += r['star']
        s['apply'] += r['apply']
        s['depts'] += 1
    for s in school_summary.values():
        s['reflow'] = s['dist_actual'] - s['dist_approved']
        s['reflow_pct'] = (round(s['reflow'] / s['dist_approved'] * 100, 1)
                           if s['dist_approved'] else None)

    # ── 機制健全性檢核：回流量原則上不應為負（缺額只回流進、不流出） ──
    neg = [(k, r) for k, r in recs.items()
           if r['reflow'] is not None and r['reflow'] < 0]
    neg_rate = len(neg) / matched * 100 if matched else 0
    if neg_rate > 10:
        print(f'⚠ 警告：{len(neg)} 個系組回流量為負（{neg_rate:.1f}%），超過10%門檻。'
              f'這通常代表欄位對應錯誤或層級彙總失敗，請覆核後再使用本輸出。')
    else:
        print(f'✔ 機制檢核：回流量為負者 {len(neg)} 個（{neg_rate:.1f}%），'
              f'多為榜單分組未能完全對應之系組，屬可接受範圍')

    print(f'✔ 比對完成：核定表 {len(recs)} 系組，其中 {matched} 個對應到榜單分組')
    print(f'  榜單系組 {len(count_rows)} 筆，已對應 {len(child_index)} 筆，'
          f'未對應 {len(unmatched)} 筆')
    if unmatched:
        print('  （未對應多為核定表未涵蓋之校系：獨立招生、境外生專班、學士後學系等）')
        for u in unmatched[:6]:
            print(f'    · {u[0]} {u[1]} (代碼{u[2]}, {u[3]}名)')

    out = {
        YEAR: recs,
        'child_index': child_index,
        'school_summary': school_summary,
        '_meta': {
            'generated_for': '116學年度落點試算',
            'year': YEAR,
            'formula': '回流量 = 回流後分發名額 - 核定分發名額；回流幅度% = 回流量 / 核定分發名額 × 100',
            'sources': {
                'approved': '教育部【表7-2】115學年度日間學士班各院系組學位學程新生招生名額分配表（公告版2026/05/06）',
                'actual': '大學考試入學分發委員會 回流後分發入學總名額（count-115.xlsx）',
            },
            'level_note': '核定表為「系」層級、榜單為「分組」層級，已用最長前綴比對將分組向上彙總後相減',
            'alloc_depts': len(recs),
            'matched_depts': matched,
            'count_rows': len(count_rows),
            'unmatched_count_rows': len(unmatched),
        },
    }
    with open(BASE / 'reflow.json', 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False)

    with open(BASE / 'reflow_unmatched.log', 'w', encoding='utf-8') as f:
        f.write('榜單校名\t榜單系組名\t代碼\t回流後名額\n')
        for u in unmatched:
            f.write('\t'.join(str(x) for x in u) + '\n')

    print('已輸出 reflow.json / reflow_unmatched.log')

    # ── 逐校回流幅度排行 ──
    print()
    print('─── 各校整體回流幅度排行（核定分發名額 ≥ 300 者）───')
    print(f'  {"學校":<18s}{"核定分發":>8s}{"回流後":>8s}{"回流量":>8s}{"回流幅度":>9s}')
    big = [(k, v) for k, v in school_summary.items() if v['dist_approved'] >= 300]
    for k, v in sorted(big, key=lambda t: -(t[1]['reflow_pct'] or 0))[:12]:
        print(f'  {k:<18s}{v["dist_approved"]:>8d}{v["dist_actual"]:>8d}'
              f'{v["reflow"]:>+8d}{v["reflow_pct"]:>8.1f}%')

    print()
    print('─── 五所重點大學 ───')
    for k in ('國立臺灣大學', '國立政治大學', '國立清華大學',
              '國立陽明交通大學', '國立成功大學'):
        v = school_summary.get(k)
        if v:
            print(f'  {k:<18s}核定{v["dist_approved"]:>5d} → 回流後{v["dist_actual"]:>5d}'
                  f'  ({v["reflow"]:+d}, {v["reflow_pct"]:+.1f}%)')

    print()
    print('─── 115學年度實戰驗證關鍵系組 ───')
    for key in ('國立臺灣大學|歷史學系', '國立臺灣大學|社會學系',
                '國立臺灣大學|生物產業傳播暨發展學系', '國立政治大學|歷史學系'):
        r = recs.get(key)
        if not r:
            print(f'  {key}: （核定表無此系組）')
            continue
        print(f'  {key}')
        print(f'      115核定小計{r["y115_total"]}（繁星{r["star"]} 申請{r["apply"]} '
              f'分發{r["dist_approved"]}）→ 回流後分發 {r["dist_actual"]} '
              f'（回流 {r["reflow"]:+d}, {r["reflow_pct"]:+.1f}%）')
        print(f'      申請占比 {r["apply_share"]}% / 分發占比 {r["dist_share"]}%'
              f' | 114核定總量 {r["y114_approved"]}')

    return 0


if __name__ == '__main__':
    sys.exit(main())
