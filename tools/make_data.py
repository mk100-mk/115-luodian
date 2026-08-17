# -*- coding: utf-8 -*-
"""
tools/make_data.py   （116學年度版，2026-08-17 改寫）

整合所有中繼資料，輸出前端用的 data/data.json。

────────────────────────────────────────────────────────────
輸入
────────────────────────────────────────────────────────────
  r115q.json    系組基本資料：採計科目/加權、學測5標檢定、英聽門檻（1,764系組）
                ※ 116學年度簡章尚未公布（約2026年11月），暫以115簡章為基底。
                  116簡章出爐後改讀 r116q.json 即可，其餘邏輯不需更動。
  hist.json     113/114/115 三年最低錄取分與錄取人數   （parse_history_results.py）
  tiebreaker.json 113/114/115 三年同分參酌門檻          （parse_tiebreaker.py）
  accu.json     113/114/115 三年官方組合成績人數累計表  （parse_accu.py）
  reflow.json   115 回流名額分析                        （parse_quota_reflow.py）
  {yr}/count-{yr}.xlsx  各年回流後分發入學總名額

────────────────────────────────────────────────────────────
相較 115 版的五項變更
────────────────────────────────────────────────────────────
1. 歷史資料由 112–114 改為 **113–115**（115已放榜，成為新的基準年）。
2. 新增**同分參酌**欄位 tb5/tb4/tb3，供前端做壓線風險警示。
3. 新增**回流名額**欄位（核定分發、回流量、回流幅度、申請/分發占比）。
4. 官方組合排名由「前端即時查表」改為**建置期同時預算各年最低分的排名/PR**。
   理由：使用者分數是即時的（仍需前端查表），但「該系最低分在當年的排名」是
   固定值，預先算好可讓前端直接顯示三年排名趨勢——而排名趨勢遠比分數趨勢穩定
   （台大歷史三年最低分 263/260/262 起伏3分，排名卻穩定在 654–706）。
5. combo 表只收 **115 年**分數區間。113/114 的 bands 不進 data.json——凡是能在
   建置期算完的（各年最低分排名、等排名換算），就不該把整張表塞給前端。

────────────────────────────────────────────────────────────
輸出 data/data.json
────────────────────────────────────────────────────────────
{
 "meta": { 版本、資料年度、方法論門檻等 },
 "rows": [ {
   "c","s","d","k"      系組代碼/校名/系組名/檢定文字
   "o"                  採計科目（簡章原始順序）
   "w"                  {科目: 加權}
   "e"                  英聽門檻(A/B)        （若有）
   "q"                  學測5標檢定條件       （若有）

   "m5","m4","m3"       115/114/113 最低錄取分
   "a5","a4","a3"       各年採計科目與加權（供前端逐年比對採計是否變動）
   "n5","n4","n3"       各年錄取人數
   "rk5","rk4","rk3"    各年最低分對應之全國組合名次（僅加權全1.00且有官方組合者）
   "pr5","pr4","pr3"    各年最低分對應之百分等級
   "tot5"               115 年該組合全國考生總人數

   "tb5","tb4","tb3"    各年同分參酌：[[科目, 門檻, 階段], ...] 依參酌順序
   "tbA"                三年皆啟動同分參酌 → true（壓線高風險）

   "st"                 115 回流後分發入學總名額
   "sa"                 115 核定登記分發名額
   "rf","rp"            回流量、回流幅度%
   "ap","dp"            申請入學占比%、分發入學占比%（核定口徑）
   "y5t","y4t"          115核定招生總量、114核定招生總量
   "st4"                114 回流後分發名額（供與 st 比較）

   "cx"                 combo 索引（僅加權全1.00且115有官方組合者）
   "eq"                 是否適用等排名換算法（115最低分 PR >= 70）
 }, ... ],
 "combo": [ {"su":[科目...], "t":總人數, "b":[[區間上限, 該分以上人數, PR], ...]} ]
}

────────────────────────────────────────────────────────────
方法論門檻（來自 parse_accu.py 的全量回測，382系組樣本）
────────────────────────────────────────────────────────────
等排名換算法在 PR>=70 時 MAE 2.2–3.5 分，明顯優於直接沿用前一年最低分（5.1–6.1分）；
但在 PR<70 反而更差（累計表尾端過平坦）。故 eq 欄位標記 PR>=70 者才適用，
前端據此決定要不要顯示等排名推估值，其餘誠實標示「無官方組合對照」。

用法：於 tools/ 目錄下執行   python make_data.py
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
OUT = BASE.parent / 'data' / 'data.json'

YEARS = ['113', '114', '115']
TAG = {'115': '5', '114': '4', '113': '3'}   # 年度 → 欄位後綴
PR_THRESHOLD = 70.0                          # 等排名換算法適用門檻（見檔頭）


def nk(school, dept):
    """跨檔案比對鍵：校名＋系組名（移除括號與空白）。

    鐵則：**禁止使用系組代碼跨年比對**——已證實 113→115 有 1,744 個代碼漂移，
    台大生傳系代碼自 0057 變為 0060，用代碼比對會取到獸醫學系。
    """
    return school + '|' + re.sub(r'[()（）\s]', '', dept)


# ───────────────────────── 載入 ─────────────────────────
def load(name):
    p = BASE / name
    if not p.exists():
        raise SystemExit(f'找不到 {p}。請先執行對應的 parse_*.py')
    return json.load(open(p, encoding='utf-8'))


r_base = load('r115q.json')
hist = load('hist.json')
tie = load('tiebreaker.json')
accu = load('accu.json')
reflow = load('reflow.json')

hist_n = {y: {nk(v['school'], v['dept']): v for v in hist[y].values()} for y in YEARS}
tie_n = {y: {nk(v['school'], v['dept']): v for v in tie[y].values()} for y in YEARS}


def load_seats(yr):
    path = BASE / yr / f'count-{yr}.xlsx'
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        school, dept, seat = r[1], r[2], r[4]
        if not school or not dept or not isinstance(seat, (int, float)):
            continue
        school = str(school).strip()
        if school.endswith('合計'):
            continue
        out[nk(school, str(dept).strip())] = int(seat)
    return out


seats = {y: load_seats(y) for y in YEARS}
print('回流後名額筆數：', {y: len(seats[y]) for y in YEARS})

# 回流資料：以榜單層級的 child_index 反查核定層級紀錄
reflow_child = reflow['child_index']
reflow_rec = reflow['115']
reflow_by_key = {}
for child_key, parent_key in reflow_child.items():
    school, dept = child_key.split('|', 1)
    reflow_by_key[nk(school, dept)] = reflow_rec[parent_key]
print('回流資料可對應之榜單系組：', len(reflow_by_key))


# ───────────────────── 官方組合查表 ─────────────────────
accu_idx = {y: {tuple(g['subjects']): g for g in accu[y]} for y in YEARS}


def rank_pr(yr, subj_codes, score):
    """該年度、該科目組合、該分數 → (名次, PR, 組合總人數)。查不到回傳 None。"""
    g = accu_idx[yr].get(tuple(sorted(subj_codes)))
    if g is None or score is None:
        return None
    for b in g['bands']:
        if score <= b['upper']:
            return b['cum_hi'], b['cum_lo_pct'], g['total']
    b = g['bands'][-1]
    return b['cum_hi'], b['cum_lo_pct'], g['total']


combo_list = []
combo_index = {}


def combo_ix(subj_tuple):
    """取得 115 年 combo 索引；該組合 115 年不存在則回傳 None。"""
    if subj_tuple in combo_index:
        return combo_index[subj_tuple]
    g = accu_idx['115'].get(subj_tuple)
    if g is None:
        return None
    idx = len(combo_list)
    combo_list.append({
        'su': list(subj_tuple),
        't': g['total'],
        'b': [[b['upper'], b['cum_hi'], b['cum_lo_pct']] for b in g['bands']],
    })
    combo_index[subj_tuple] = idx
    return idx


def all_one(weights):
    return bool(weights) and all(abs(v - 1.0) < 1e-9 for v in weights.values())


# ───────────────────────── 逐系組整合 ─────────────────────────
rows = []
stat = {k: 0 for k in ('m5', 'm4', 'm3', 'tb5', 'tbA', 'st', 'rf', 'cx', 'eq', 'rk5')}

for r in r_base:
    key = nk(r['school'], r['dept'])
    row = {
        'c': r['code'], 's': r['school'], 'd': r['dept'], 'k': r['check'],
        'o': [x[0] for x in sorted(r['subjects'], key=lambda t: (t[2] or 9))],
        'w': {x[0]: x[1] for x in r['subjects']},
    }
    m = re.search(r'英\s*聽\s*[（(]?\s*([ABＡＢ])\s*級', r['check'])
    if m:
        row['e'] = {'Ａ': 'A', 'Ｂ': 'B'}.get(m.group(1), m.group(1))
    if r.get('q'):
        row['q'] = [[[s, d] for s, d in cl] for cl in r['q']]

    # ── 三年最低分 / 錄取人數 / 採計加權 / 排名 ──
    for y in YEARS:
        t = TAG[y]
        h = hist_n[y].get(key)
        if not h:
            continue
        score = float(h['score']) if h['score'] else None
        row['m' + t] = score
        row['a' + t] = h['subjects']
        row['n' + t] = h['n']
        if score is not None:
            stat['m' + t] = stat.get('m' + t, 0) + 1
        # 排名/PR 僅在「該年採計加權全為1.00且官方有該組合」時才成立
        if score is not None and all_one(h['subjects']):
            rp = rank_pr(y, list(h['subjects'].keys()), score)
            if rp:
                row['rk' + t], row['pr' + t] = rp[0], rp[1]
                if y == '115':
                    row['tot5'] = rp[2]
                    stat['rk5'] += 1

    # ── 三年同分參酌 ──
    hit_years = 0
    for y in YEARS:
        tb = tie_n[y].get(key)
        if not tb:
            continue
        hit_years += 1
        row['tb' + TAG[y]] = [[lv['subject'], lv['min'], lv['stage']] for lv in tb['levels']]
    if hit_years == len(YEARS):
        row['tbA'] = True
        stat['tbA'] += 1
    if 'tb5' in row:
        stat['tb5'] += 1

    # ── 名額（回流後） ──
    if key in seats['115']:
        row['st'] = seats['115'][key]
        stat['st'] += 1
    if key in seats['114']:
        row['st4'] = seats['114'][key]

    # ── 回流分析 ──
    rf = reflow_by_key.get(key)
    if rf:
        row['sa'] = rf['dist_approved']
        row['rf'] = rf['reflow']
        row['rp'] = rf['reflow_pct']
        row['ap'] = rf['apply_share']
        row['dp'] = rf['dist_share']
        row['y5t'] = rf['y115_total']
        row['y4t'] = rf['y114_approved']
        stat['rf'] += 1

    # ── 官方組合排名（依本年度簡章之採計加權判定） ──
    if all_one(row['w']):
        cx = combo_ix(tuple(sorted(row['w'].keys())))
        if cx is not None:
            row['cx'] = cx
            stat['cx'] += 1
            # 等排名換算法適用性：以該系 115 最低分之 PR 判定
            if row.get('pr5') is not None and row['pr5'] >= PR_THRESHOLD:
                row['eq'] = True
                stat['eq'] += 1

    rows.append(row)

output = {
    'meta': {
        'version': '116.1',
        'generated_for': '116學年度分科分發落點試算',
        'base_prospectus': '115學年度簡章（116簡章未公布前之暫用基底）',
        'history_years': YEARS,
        'baseline_year': '115',
        'join_key': '校名+系組名（禁止使用系組代碼跨年比對）',
        'pr_threshold_for_equating': PR_THRESHOLD,
        'tiebreak_warn_window': 2,
        'notes': [
            '總分達標不等於錄取：與最低分差距在±2分內時，須檢視同分參酌門檻',
            '分發名額＝核定名額＋申請入學缺額回流，全國回流幅度中位數約+48%',
            '官方組合排名僅適用加權全1.00且官方公布對照組別者，其餘以歷年比較法為輔',
            '排名趨勢比分數趨勢穩定，優先參考 rk3/rk4/rk5',
        ],
    },
    'rows': rows,
    'combo': combo_list,
}

OUT.parent.mkdir(parents=True, exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

# ───────────────────────── 統計摘要 ─────────────────────────
size_kb = OUT.stat().st_size / 1024
print()
print(f'已輸出 {OUT}（{size_kb:,.0f} KB）')
print(f'系組數: {len(rows)}   官方組合表(去重後): {len(combo_list)}')
print()
print(f'{"欄位":<28s}{"系組數":>7s}{"覆蓋率":>9s}')
cov = [
    ('115最低分 (m5)', stat['m5']), ('114最低分 (m4)', stat['m4']),
    ('113最低分 (m3)', stat['m3']),
    ('115最低分之全國排名 (rk5)', stat['rk5']),
    ('115同分參酌 (tb5)', stat['tb5']),
    ('三年皆同分參酌 (tbA)', stat['tbA']),
    ('115回流後名額 (st)', stat['st']),
    ('回流分析 (rf)', stat['rf']),
    ('官方組合排名 (cx)', stat['cx']),
    ('適用等排名換算 (eq)', stat['eq']),
]
for name, n in cov:
    print(f'{name:<28s}{n:>7d}{n / len(rows) * 100:>8.1f}%')

# ───────────────────────── 抽樣核對 ─────────────────────────
print()
print('─── 抽樣核對：115學年度實戰驗證關鍵系組 ───')
for name in ('國立臺灣大學|歷史學系', '國立臺灣大學|社會學系',
             '國立臺灣大學|生物產業傳播暨發展學系'):
    row = next((x for x in rows if nk(x['s'], x['d']) == name), None)
    if not row:
        print(f'\n{name}: 找不到')
        continue
    print(f'\n{name}  （代碼 {row["c"]}）')
    print('  年度   最低分   錄取   全國名次      PR   同分參酌')
    for y in ('115', '114', '113'):
        t = TAG[y]
        if 'm' + t not in row:
            print(f'  {y}    （無資料）')
            continue
        tb = row.get('tb' + t)
        tbs = ' → '.join(f'{a}≥{b:g}' for a, b, _ in tb) if tb else '未啟動'
        rk = row.get('rk' + t)
        pr = row.get('pr' + t)
        print(f'  {y}   {row["m" + t]:>7} {row.get("n" + t, "-"):>5}   '
              f'{rk if rk is not None else "-":>8}  {pr if pr is not None else "-":>6}   {tbs}')
    if 'rf' in row:
        print(f'  名額：核定分發 {row["sa"]} → 回流後 {row["st"]} '
              f'（回流 {row["rf"]:+d}, {row["rp"]:+.1f}%）｜'
              f'申請占比 {row["ap"]}% / 分發占比 {row["dp"]}%')
    print(f'  官方組合對照：{"有" if "cx" in row else "無（加權非1.00或無對照組別）"}'
          f'｜等排名換算適用：{"是" if row.get("eq") else "否"}'
          f'｜三年皆壓線：{"是" if row.get("tbA") else "否"}')

sys.exit(0)
